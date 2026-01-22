"""
Excel report generator with 4 sheets and conditional formatting.

Sheet 1: Metadata Summary
Sheet 2: Digital Signatures
Sheet 3: Form Fields
Sheet 4: Extraction Quality
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..core.extraction_result import PDFMetadataResult, ExecutionStatus

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates 4-sheet Excel reports with conditional formatting"""

    # Colors for conditional formatting
    COLOR_GREEN = "C6EFCE"      # Executed, high confidence
    COLOR_YELLOW = "FFEB9C"     # Low confidence
    COLOR_GRAY = "D9D9D9"       # Not executed
    COLOR_RED = "FFC7CE"        # Errors

    def __init__(self):
        """Initialize Excel generator"""
        pass

    def generate(
        self,
        results: List[PDFMetadataResult],
        output_path: Path,
        include_timestamp: bool = True
    ):
        """
        Generate Excel report.

        Args:
            results: List of extraction results
            output_path: Output file path
            include_timestamp: Include timestamp in filename
        """
        logger.info(f"Generating Excel report for {len(results)} PDFs...")

        # Add timestamp to filename if requested
        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            stem = output_path.stem
            output_path = output_path.parent / f"{stem}_{timestamp}{output_path.suffix}"

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets
        self._create_metadata_summary_sheet(wb, results)
        self._create_signatures_sheet(wb, results)
        self._create_form_fields_sheet(wb, results)
        self._create_quality_sheet(wb, results)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")

    def _create_metadata_summary_sheet(self, wb: Workbook, results: List[PDFMetadataResult]):
        """Create Sheet 1: Metadata Summary"""
        ws = wb.create_sheet("Metadata Summary", 0)

        # Headers
        headers = [
            "Filename", "Creation Date", "Mod Date", "Author", "Producer",
            "Pages", "File Size (MB)", "Encrypted", "Has Signatures",
            "Is Executed", "Execution Confidence", "Overall Confidence", "Notes"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Write data
        for row_idx, result in enumerate(results, 2):
            meta = result.document_metadata

            ws.cell(row_idx, 1, meta.filename)
            ws.cell(row_idx, 2, meta.creation_date.strftime('%Y-%m-%d') if meta.creation_date else "")
            ws.cell(row_idx, 3, meta.modification_date.strftime('%Y-%m-%d') if meta.modification_date else "")
            ws.cell(row_idx, 4, meta.author or "")
            ws.cell(row_idx, 5, meta.producer or "")
            ws.cell(row_idx, 6, meta.page_count)
            ws.cell(row_idx, 7, f"{meta.file_size_bytes / (1024*1024):.2f}")
            ws.cell(row_idx, 8, "Yes" if meta.is_encrypted else "No")
            ws.cell(row_idx, 9, "Yes" if result.has_digital_signatures else "No")
            ws.cell(row_idx, 10, result.execution_status.value)
            ws.cell(row_idx, 11, result.execution_confidence)
            ws.cell(row_idx, 12, result.overall_confidence)
            ws.cell(row_idx, 13, meta.notes)

            # Conditional formatting based on execution status
            status_cell = ws.cell(row_idx, 10)
            confidence_cell = ws.cell(row_idx, 11)

            if result.execution_status == ExecutionStatus.FULLY_EXECUTED:
                if result.execution_confidence >= 80:
                    status_cell.fill = PatternFill(start_color=self.COLOR_GREEN, fill_type="solid")
                elif result.execution_confidence < 50:
                    status_cell.fill = PatternFill(start_color=self.COLOR_YELLOW, fill_type="solid")
            elif result.execution_status == ExecutionStatus.UNSIGNED:
                status_cell.fill = PatternFill(start_color=self.COLOR_GRAY, fill_type="solid")

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        ws.column_dimensions['A'].width = 40  # Filename wider

    def _create_signatures_sheet(self, wb: Workbook, results: List[PDFMetadataResult]):
        """Create Sheet 2: Digital Signatures"""
        ws = wb.create_sheet("Digital Signatures", 1)

        # Headers
        headers = [
            "Filename", "Signature Name", "Signer", "Signing Time",
            "Signature Type", "Certificate Issuer", "Valid From", "Valid To",
            "Is Valid", "Covers Document", "Confidence", "Notes"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Write data
        row_idx = 2
        for result in results:
            if not result.digital_signatures:
                # Still show filename even if no signatures
                ws.cell(row_idx, 1, result.document_metadata.filename)
                ws.cell(row_idx, 13, "No signatures found")
                row_idx += 1
            else:
                for sig in result.digital_signatures:
                    ws.cell(row_idx, 1, result.document_metadata.filename)
                    ws.cell(row_idx, 2, sig.signature_name)
                    ws.cell(row_idx, 3, sig.signer_name or "")
                    ws.cell(row_idx, 4, sig.signing_time.strftime('%Y-%m-%d %H:%M') if sig.signing_time else "")
                    ws.cell(row_idx, 5, sig.signature_type)
                    ws.cell(row_idx, 6, sig.certificate_issuer or "")
                    ws.cell(row_idx, 7, sig.certificate_valid_from.strftime('%Y-%m-%d') if sig.certificate_valid_from else "")
                    ws.cell(row_idx, 8, sig.certificate_valid_to.strftime('%Y-%m-%d') if sig.certificate_valid_to else "")
                    ws.cell(row_idx, 9, "Yes" if sig.is_valid is True else ("No" if sig.is_valid is False else "Unknown"))
                    ws.cell(row_idx, 10, "Yes" if sig.signature_covers_document else "No")
                    ws.cell(row_idx, 11, sig.confidence)
                    ws.cell(row_idx, 12, sig.notes)

                    row_idx += 1

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        ws.column_dimensions['A'].width = 40

    def _create_form_fields_sheet(self, wb: Workbook, results: List[PDFMetadataResult]):
        """Create Sheet 3: Form Fields"""
        ws = wb.create_sheet("Form Fields", 2)

        # Headers
        headers = [
            "Filename", "Field Name", "Field Type", "Field Value",
            "Required", "Read Only", "Page Number"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Write data
        row_idx = 2
        for result in results:
            if not result.form_fields:
                ws.cell(row_idx, 1, result.document_metadata.filename)
                ws.cell(row_idx, 7, "No form fields")
                row_idx += 1
            else:
                for field in result.form_fields:
                    ws.cell(row_idx, 1, result.document_metadata.filename)
                    ws.cell(row_idx, 2, field.field_name)
                    ws.cell(row_idx, 3, field.field_type)
                    ws.cell(row_idx, 4, field.field_value or "")
                    ws.cell(row_idx, 5, "Yes" if field.is_required else "No")
                    ws.cell(row_idx, 6, "Yes" if field.is_read_only else "No")
                    ws.cell(row_idx, 7, field.page_number)

                    row_idx += 1

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        ws.column_dimensions['A'].width = 40

    def _create_quality_sheet(self, wb: Workbook, results: List[PDFMetadataResult]):
        """Create Sheet 4: Extraction Quality"""
        ws = wb.create_sheet("Extraction Quality", 3)

        # Calculate metrics
        total_pdfs = len(results)
        pdfs_with_sigs = sum(1 for r in results if r.has_digital_signatures)
        pdfs_with_valid_sigs = sum(1 for r in results if r.has_valid_signatures)
        executed_contracts = sum(1 for r in results if r.execution_status == ExecutionStatus.FULLY_EXECUTED)
        high_confidence = sum(1 for r in results if r.overall_confidence >= 70)
        extraction_errors = sum(1 for r in results if r.error is not None)
        avg_processing_time = sum(r.processing_time_seconds for r in results) / total_pdfs if total_pdfs > 0 else 0

        # Metrics
        metrics = [
            ("Total PDFs Processed", total_pdfs),
            ("PDFs with Digital Signatures", f"{pdfs_with_sigs} ({pdfs_with_sigs/total_pdfs*100:.1f}%)" if total_pdfs > 0 else "0"),
            ("PDFs with Valid Signatures", f"{pdfs_with_valid_sigs} ({pdfs_with_valid_sigs/total_pdfs*100:.1f}%)" if total_pdfs > 0 else "0"),
            ("Executed Contracts", f"{executed_contracts} ({executed_contracts/total_pdfs*100:.1f}%)" if total_pdfs > 0 else "0"),
            ("High Confidence Extractions (≥70%)", f"{high_confidence} ({high_confidence/total_pdfs*100:.1f}%)" if total_pdfs > 0 else "0"),
            ("Extraction Errors", extraction_errors),
            ("Average Processing Time", f"{avg_processing_time:.2f}s"),
        ]

        # Write headers
        ws.cell(1, 1, "Metric").font = Font(bold=True)
        ws.cell(1, 2, "Value").font = Font(bold=True)

        # Write metrics
        for row_idx, (metric, value) in enumerate(metrics, 2):
            ws.cell(row_idx, 1, metric)
            ws.cell(row_idx, 2, value)

        # Auto-size columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        logger.info(f"Quality metrics: {executed_contracts}/{total_pdfs} executed, "
                   f"{high_confidence}/{total_pdfs} high confidence")
