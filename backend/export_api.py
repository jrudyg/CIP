"""
CIP Data Export API

Provides endpoints for exporting contract data:
- /api/export/contracts - Export contracts as CSV, JSON, or Excel
- /api/export/summary - Export summary report

Supports filtering by status, type, risk level, and date range.
"""

import csv
import io
import json
import sqlite3
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional

from flask import Blueprint, request, Response, jsonify

# Create blueprint
export_bp = Blueprint('export', __name__, url_prefix='/api/export')

# Configuration
CIP_ROOT = Path(r"C:\Users\jrudy\CIP")
CONTRACTS_DB = CIP_ROOT / "data" / "contracts.db"

# Exportable fields with display names
EXPORT_FIELDS = {
    'id': 'ID',
    'title': 'Title',
    'filename': 'Filename',
    'counterparty': 'Counterparty',
    'contract_type': 'Contract Type',
    'status': 'Status',
    'risk_level': 'Risk Level',
    'contract_value': 'Contract Value',
    'effective_date': 'Effective Date',
    'expiration_date': 'Expiration Date',
    'purpose': 'Purpose',
    'upload_date': 'Upload Date',
    'created_at': 'Created At',
    'updated_at': 'Updated At',
}

DEFAULT_FIELDS = [
    'id', 'title', 'counterparty', 'contract_type', 'status',
    'risk_level', 'contract_value', 'effective_date', 'expiration_date'
]


def get_db_connection():
    """Get database connection with row factory."""
    conn = sqlite3.connect(str(CONTRACTS_DB))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def build_export_query(
    fields: List[str],
    status: Optional[List[str]] = None,
    contract_type: Optional[List[str]] = None,
    risk_level: Optional[List[str]] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_archived: bool = False
) -> tuple:
    """Build SQL query with filters."""

    # Validate fields
    valid_fields = [f for f in fields if f in EXPORT_FIELDS]
    if not valid_fields:
        valid_fields = DEFAULT_FIELDS

    select_clause = ", ".join(valid_fields)
    where_clauses = []
    params = []

    # Base filter: exclude archived unless requested
    if not include_archived:
        where_clauses.append("(archived = 0 OR archived IS NULL)")

    # Status filter
    if status:
        placeholders = ", ".join("?" * len(status))
        where_clauses.append(f"status IN ({placeholders})")
        params.extend(status)

    # Contract type filter
    if contract_type:
        placeholders = ", ".join("?" * len(contract_type))
        where_clauses.append(f"contract_type IN ({placeholders})")
        params.extend(contract_type)

    # Risk level filter
    if risk_level:
        placeholders = ", ".join("?" * len(risk_level))
        where_clauses.append(f"risk_level IN ({placeholders})")
        params.extend(risk_level)

    # Date range filter (on expiration_date)
    if date_from:
        where_clauses.append("expiration_date >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("expiration_date <= ?")
        params.append(date_to)

    where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"

    query = f"""
        SELECT {select_clause}
        FROM contracts
        WHERE {where_clause}
        ORDER BY id DESC
    """

    return query, params, valid_fields


def format_value(value: Any, field: str) -> str:
    """Format a value for export."""
    if value is None:
        return ""
    if field == 'contract_value':
        return f"{value:.2f}" if isinstance(value, (int, float)) else str(value)
    return str(value)


def export_to_csv(rows: List[Dict], fields: List[str]) -> str:
    """Export data to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row with display names
    headers = [EXPORT_FIELDS.get(f, f) for f in fields]
    writer.writerow(headers)

    # Data rows
    for row in rows:
        writer.writerow([format_value(row.get(f), f) for f in fields])

    return output.getvalue()


def export_to_json(rows: List[Dict], fields: List[str]) -> str:
    """Export data to JSON string."""
    # Filter to requested fields only
    filtered_rows = [
        {f: row.get(f) for f in fields}
        for row in rows
    ]
    return json.dumps({
        'contracts': filtered_rows,
        'count': len(filtered_rows),
        'fields': fields,
        'exported_at': datetime.now().isoformat()
    }, indent=2, default=str)


def export_to_excel(rows: List[Dict], fields: List[str]) -> bytes:
    """Export data to Excel bytes (using CSV as fallback if openpyxl not available)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Contracts"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        # Headers
        headers = [EXPORT_FIELDS.get(f, f) for f in fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row.get(field)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, 1):
            max_length = len(EXPORT_FIELDS.get(field, field))
            for row in rows[:100]:  # Sample first 100 rows
                val = str(row.get(field, ""))
                max_length = max(max_length, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return export_to_csv(rows, fields).encode('utf-8')


@export_bp.route('/contracts', methods=['GET'])
def export_contracts():
    """
    Export contracts data.

    Query params:
        - format: csv|json|xlsx (default: csv)
        - status: comma-separated list (e.g., active,negotiating)
        - type: comma-separated contract types (e.g., MSA,SOW)
        - risk: comma-separated risk levels (e.g., HIGH,CRITICAL)
        - from: start date YYYY-MM-DD (filters expiration_date)
        - to: end date YYYY-MM-DD (filters expiration_date)
        - fields: comma-separated field names (default: core fields)
        - include_archived: true to include archived contracts

    Returns:
        File download in requested format
    """
    # Parse parameters
    export_format = request.args.get('format', 'csv').lower()
    status = request.args.get('status', '').split(',') if request.args.get('status') else None
    contract_type = request.args.get('type', '').split(',') if request.args.get('type') else None
    risk_level = request.args.get('risk', '').split(',') if request.args.get('risk') else None
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    fields = request.args.get('fields', '').split(',') if request.args.get('fields') else DEFAULT_FIELDS
    include_archived = request.args.get('include_archived', 'false').lower() == 'true'

    # Clean up empty strings from lists
    if status:
        status = [s.strip() for s in status if s.strip()]
    if contract_type:
        contract_type = [t.strip() for t in contract_type if t.strip()]
    if risk_level:
        risk_level = [r.strip().upper() for r in risk_level if r.strip()]
    if fields:
        fields = [f.strip() for f in fields if f.strip()]

    # Build and execute query
    query, params, valid_fields = build_export_query(
        fields=fields,
        status=status,
        contract_type=contract_type,
        risk_level=risk_level,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived
    )

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Export based on format
    if export_format == 'json':
        content = export_to_json(rows, valid_fields)
        return Response(
            content,
            mimetype='application/json',
            headers={
                'Content-Disposition': f'attachment; filename=contracts_export_{timestamp}.json'
            }
        )

    elif export_format in ('xlsx', 'excel'):
        content = export_to_excel(rows, valid_fields)
        return Response(
            content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=contracts_export_{timestamp}.xlsx'
            }
        )

    else:  # Default to CSV
        content = export_to_csv(rows, valid_fields)
        return Response(
            content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=contracts_export_{timestamp}.csv'
            }
        )


@export_bp.route('/contracts/preview', methods=['GET'])
def preview_export():
    """
    Preview export data without downloading.

    Same parameters as /contracts endpoint.
    Returns JSON preview with first 10 rows.
    """
    # Parse parameters (same as export)
    status = request.args.get('status', '').split(',') if request.args.get('status') else None
    contract_type = request.args.get('type', '').split(',') if request.args.get('type') else None
    risk_level = request.args.get('risk', '').split(',') if request.args.get('risk') else None
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    fields = request.args.get('fields', '').split(',') if request.args.get('fields') else DEFAULT_FIELDS
    include_archived = request.args.get('include_archived', 'false').lower() == 'true'

    # Clean up
    if status:
        status = [s.strip() for s in status if s.strip()]
    if contract_type:
        contract_type = [t.strip() for t in contract_type if t.strip()]
    if risk_level:
        risk_level = [r.strip().upper() for r in risk_level if r.strip()]
    if fields:
        fields = [f.strip() for f in fields if f.strip()]

    # Build query
    query, params, valid_fields = build_export_query(
        fields=fields,
        status=status,
        contract_type=contract_type,
        risk_level=risk_level,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived
    )

    conn = get_db_connection()
    cursor = conn.cursor()

    # Get total count
    count_query = query.replace(f"SELECT {', '.join(valid_fields)}", "SELECT COUNT(*)")
    count_query = count_query.split("ORDER BY")[0]
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]

    # Get preview rows
    cursor.execute(query + " LIMIT 10", params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify({
        'preview': rows,
        'total_count': total_count,
        'fields': valid_fields,
        'field_names': {f: EXPORT_FIELDS.get(f, f) for f in valid_fields},
        'filters_applied': {
            'status': status,
            'contract_type': contract_type,
            'risk_level': risk_level,
            'date_from': date_from,
            'date_to': date_to,
            'include_archived': include_archived
        }
    })


@export_bp.route('/fields', methods=['GET'])
def list_export_fields():
    """
    List available export fields.

    Returns:
        {fields: {field_name: display_name}, default_fields: [...]}
    """
    return jsonify({
        'fields': EXPORT_FIELDS,
        'default_fields': DEFAULT_FIELDS
    })


@export_bp.route('/summary', methods=['GET'])
def export_summary():
    """
    Export a summary report.

    Returns JSON with portfolio statistics suitable for reporting.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    base_where = "(archived = 0 OR archived IS NULL)"

    summary = {
        'generated_at': datetime.now().isoformat(),
        'portfolio': {},
        'by_status': [],
        'by_type': [],
        'by_risk': [],
        'expiring_soon': []
    }

    # Portfolio totals
    cursor.execute(f"SELECT COUNT(*) FROM contracts WHERE {base_where}")
    summary['portfolio']['total_contracts'] = cursor.fetchone()[0]

    cursor.execute(f"SELECT COALESCE(SUM(contract_value), 0) FROM contracts WHERE {base_where}")
    summary['portfolio']['total_value'] = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM contracts WHERE {base_where} AND status = 'active'")
    summary['portfolio']['active_contracts'] = cursor.fetchone()[0]

    # By status
    cursor.execute(f"""
        SELECT status, COUNT(*) as count, COALESCE(SUM(contract_value), 0) as value
        FROM contracts WHERE {base_where} AND status IS NOT NULL
        GROUP BY status ORDER BY count DESC
    """)
    for row in cursor.fetchall():
        summary['by_status'].append({
            'status': row['status'],
            'count': row['count'],
            'value': row['value']
        })

    # By type
    cursor.execute(f"""
        SELECT contract_type, COUNT(*) as count, COALESCE(SUM(contract_value), 0) as value
        FROM contracts WHERE {base_where} AND contract_type IS NOT NULL
        GROUP BY contract_type ORDER BY count DESC
    """)
    for row in cursor.fetchall():
        summary['by_type'].append({
            'type': row['contract_type'],
            'count': row['count'],
            'value': row['value']
        })

    # By risk
    cursor.execute(f"""
        SELECT risk_level, COUNT(*) as count, COALESCE(SUM(contract_value), 0) as value
        FROM contracts WHERE {base_where} AND risk_level IS NOT NULL
        GROUP BY risk_level
        ORDER BY CASE risk_level WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END
    """)
    for row in cursor.fetchall():
        summary['by_risk'].append({
            'risk_level': row['risk_level'],
            'count': row['count'],
            'value': row['value']
        })

    # Expiring in next 90 days
    from datetime import timedelta
    today = datetime.now().date().isoformat()
    d90 = (datetime.now().date() + timedelta(days=90)).isoformat()

    cursor.execute(f"""
        SELECT id, title, counterparty, expiration_date, contract_value, risk_level
        FROM contracts
        WHERE {base_where}
        AND expiration_date BETWEEN ? AND ?
        AND status NOT IN ('expired', 'terminated')
        ORDER BY expiration_date
    """, (today, d90))

    for row in cursor.fetchall():
        summary['expiring_soon'].append(dict(row))

    conn.close()

    return jsonify(summary)


def register_export(app):
    """Register export blueprint with Flask app."""
    app.register_blueprint(export_bp)
    return export_bp
